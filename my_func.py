import os
import smtplib
import time
from selenium import webdriver
import pandas as pd
import xlsxwriter
import openpyxl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from bash import bash
from openpyxl.styles import Font, Border, Side, Style, Color, PatternFill, Alignment
from openpyxl.utils import coordinate_from_string

PATH=os.getcwd()

def merge_df_and_save_to_csv(dict_with_url, number_of_progons):
    columns_to_skip = ['Package','Class']
    timestr = time.strftime("%Y.%m.%d")
    excel_writer = pd.ExcelWriter("report/"+timestr+".xlsx", engine='xlsxwriter')
    for operator in dict_with_url:
        for i in range (1,len(operator['url'])+1):
            df = pd.read_csv("download/{}{}.csv".format(operator['provider'],i),usecols=lambda x: x not in columns_to_skip)
            df = remove_rows_csv(df,number_of_progons,["SORM.test","start_sorm","test_stop"])
            if i==1:
                header=list(df.columns)
                finaldf=pd.DataFrame(columns=header)
            finaldf=finaldf.append(df,sort=False)
        finaldf.to_csv("download/"+operator['provider']+".csv", index=False)
    return excel_writer, timestr

def add_summary_to_df(df):
    passed=countif_in_rows("PASSED",df)
    failed=countif_in_rows("FAILED",df)
    skipped=countif_in_rows("SKIPPED",df)
    results=["SUMMARY:"]
    for i in range (0,len(passed)-1):
        results.append(" ")
    df.loc[len(df)]=results
    df.loc[len(df)]=passed
    df.loc[len(df)]=failed
    df.loc[len(df)]=skipped      
    return df,passed,failed,skipped

def get_url(operators):    
    for operator in operators:
        download_url=[]
        provider=operator['provider']
        dvo_list=operator['dvo']
        job=operator['job']
        for dvo in dvo_list:
            if operator['provider']=="ALL_OPERATORS_IMS":
                download_url.append("http://10.72.1.46/view/{provider}/job/{job}/OPERATOR={dvo},label=SORM_server/test_results_analyzer/".format(provider=provider,job=job, dvo=dvo))
            else:
                download_url.append("http://10.72.1.46/view/{provider}/job/{job}/dumptemplate={dvo},label=SORM_server/test_results_analyzer/".format(provider=provider,job=job, dvo=dvo))
        operator['url']=download_url
    return operators

def download_from_url(url_list,num,sleep=3):
    try:
        if os.path.exists("download"):
            bash("rm -rf ./download")
        if not os.path.exists("report"):
            os.mkdir("report")
        os.mkdir("download")
        options = webdriver.ChromeOptions()
        options.headless = True
        # options.binary_location="/usr/bin/chromium-browser"
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        driver = webdriver.Chrome(PATH+'/chromedriver',chrome_options=options)  # Optional argument, if not specified will search path.
        driver.command_executor._commands["send_command"] = ("POST", '/session/$sessionId/chromium/send_command')
        params = {'cmd': 'Page.setDownloadBehavior', 'params': {'behavior': 'allow', 'downloadPath': PATH+"/download/"}}
        command_result = driver.execute("send_command", params)
        for operator in url_list:
            n=len(operator['url'])
            for url in operator['url']:
                newname="download/{}{}.csv".format(operator['provider'],str(n))
                driver.get(url);
                if num!=5:
                    search_box = driver.find_element_by_id("settingsmenubutton").click()
                    driver.find_element_by_id("noofbuilds").clear()
                    driver.find_element_by_id("noofbuilds").send_keys(str(num))
                    time.sleep(sleep)
                search_box = driver.find_element_by_id('downloadCSV').click()
                time.sleep(1.5)
                os.rename('download/Test Results.csv', newname)
                n-=1
        time.sleep(2)
    finally:
    	print("Скачалось")
    return True

def remove_rows_csv(df,num,word_list=["Esli eto slovo est v stroke, to stroka udalit'sya"]):
    #if row will contain word, row will be deleted
    x=[]
    for index, row in df.iterrows():
        n=0
        for item in row:
            if type(item)==float:
                n+=1
            if type(item)==str:
                for word in word_list:
                    if item.count(word)>0:
                        x.append(index)
                        continue
                if item.count("\\")>2:
                    x.append(index)
                    continue
        if n==num:
            x.append(index)
    df1=df.drop(x, axis=0)
    return df1

def num_to_letter(n):
    ###Example of this function: input=6, output=F
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

def rename_test_column(path):
    df = pd.read_csv(path)
    for i in range(0,len(df)):
        df["Test"][i]=df["Test"][i][14:-5]
    ###Merge rows with NA values
    df=df.fillna('').groupby('Test',as_index=False).agg('sum')
    df.to_csv(path,index=False)
    return(df)

def set_border(ws, cell_range, fill=False, color="DDD9C4"):
    border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))
    if fill==False:
        tmp=cell_range.split(":")
        a=tmp[0][1:]
        b=tmp[1][0]+a
        header=tmp[0]+":"+b
    
        rows = ws[cell_range]
        for row in rows:
            for cell in row:
                cell.border = border
    
        head=ws[header]
        for row in head:
            for cell in row:
                cell.style=Style(fill=PatternFill(patternType='solid',
                                        fill_type='solid', 
                                        fgColor=Color(color)))
                cell.border = border
    if fill==True:
        rows = ws[cell_range]
        for row in rows:
            for cell in row:
                cell.style=Style(fill=PatternFill(patternType='solid',
                                        fill_type='solid', 
                                        fgColor=Color(color)))
                cell.border = border

def countif_in_rows(word,df):
##count in each columns by word, return list of counts
    x=['TOTAL '+word]
    n=0
    for column in df.columns[1:]:
        for cell in df[column]:
            if cell==word:
                n=n+1
        x.append(n)
        n=0
    return(x)

def create_diff_table(df,provider,builds_to_diff="default",include_provider_name=False):
    if builds_to_diff=="default":
        build1=1
        build2=2
    else:
        try:
            build1=list(df.columns).index(str(builds_to_diff[0]))
            build2=list(df.columns).index(str(builds_to_diff[1]))
        except ValueError:
        	print("В отчете нет информации по сборкам, которые хотите сравнить")
        	empty_df=pd.DataFrame({'A' : []})
        	return empty_df

    # df=pd.read_csv(csv)
    test=df[df.columns[0]]
    today=df[df.columns[build1]]
    yesterday=df[df.columns[build2]]
    
    if include_provider_name:
        passed=[" "]
        failed=[provider]
        norun=[" "]
    else:
        passed=[]
        failed=[]
        norun=[]    	
    for i in range(0,len(test)-1):
        if str(today[i])==str(yesterday[i]):
            continue
        if str(today[i])=='PASSED' and str(yesterday)!="PASSED":
            passed.append(test[i])
        if str(today[i])=='FAILED' and (str(yesterday)!='FAILED' or str(yesterday)=='SKIPPED'):
            failed.append(test[i])
        if str(today[i])=='' and str(yesterday)!='':
            norun.append(test[i])
    
    s1=pd.Series(passed,name='Стали работать')
    s2=pd.Series(failed,name='Перестали работать')
    s3=pd.Series(norun,name='Перестали запускаться')
    
    
    df1=pd.concat([s1,s2,s3], axis=1)
    return df1

def get_maximum_row(ws, column):
    try:
        return max(coordinate_from_string(cell)[-1]
            for cell in ws._cells if cell.startswith(column))
    except ValueError:
        return 1

def append_df_to_excel(filename, df, startcol, sheet_name='Sheet1', descr=None,
	                   startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook


        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        # if startrow is None and sheet_name in writer.book.sheetnames:
        #     startrow = writer.book[sheet_name].max_row

        startrow=max(get_maximum_row(writer.book[sheet_name],num_to_letter(startcol+1)),
                     get_maximum_row(writer.book[sheet_name],num_to_letter(startcol+2)),
                     get_maximum_row(writer.book[sheet_name],num_to_letter(startcol+3)))

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow+2, startcol=startcol, **to_excel_kwargs)
    if descr!=None:
        writer.book[sheet_name].merge_cells(num_to_letter(startcol+1)+str(startrow+2)+":"+num_to_letter(startcol+3)+str(startrow+2))
        writer.book[sheet_name][num_to_letter(startcol+1)+str(startrow+2)]=descr
        writer.book[sheet_name].cell(num_to_letter(startcol+1)+str(startrow+2)).alignment = Alignment(horizontal='center')
        writer.book[sheet_name].cell(num_to_letter(startcol+1)+str(startrow+2)).font = Font(color="FF0000", bold=True)
    set_border(writer.book[sheet_name], num_to_letter(startcol+1)+str(startrow+3)+":"+num_to_letter(startcol+3)+str(startrow+3+len(df)))
    # save the workbook
    writer.save()

def add_user_diff_to_excel(builds_to_diff, timestr, real_num_of_progons):
    try:
        df=pd.read_csv("download/"+builds_to_diff['provider']+".csv")
        user_diff_table=create_diff_table(df,builds_to_diff['provider'],builds_to_diff=builds_to_diff['builds'])
        if not user_diff_table.empty:
            descr="Таблица сравнения сборок {} и {}".format(builds_to_diff['builds'][0],builds_to_diff['builds'][1])
            append_df_to_excel("report/"+timestr+".xlsx", user_diff_table, startcol=real_num_of_progons+2, sheet_name=builds_to_diff['provider'],descr=descr, index=False)
    except FileNotFoundError:
        print("Скорее всего неправильный builds_to_diff, таблицы сравнения указанных версий в отчете не будет")
        return None
def get_version_df(operator,header):
    try:
        os.mkdir("download/results")
        ver_df=pd.DataFrame(columns=["component"])
        for id in header[1:3]:
            result_url="//10.72.1.239/Results/{job}/{id}".format(job=operator['result'],id=id)
            if operator["provider"]=="ALL_OPERATORS_IMS":
                bash("wget -P ./download/results ftp:"+ result_url+"/Beeline/result.tar.xz")
            else:
                bash("wget -P ./download/results ftp:"+ result_url+"/CALLDIR/result.tar.xz")
            bash("tar xf ./download/results/result.tar.xz -C ./download/results/")
            f=open("./download/results/versions.log","r")
            versions=f.read()
            f.close()  

            keys=[]
            values=[]
            versions_list=versions.split("\n")
            for item in versions_list[:-1]:
                keys.append(item.split("/")[0])
                values.append(item.split("/")[1])
            ver_df['component']=keys
            ver_df[id]=values

    except FileNotFoundError:
            keys=[" "]
            values=["Нет информации, возможно тест не запустился"]
            ver_df['component']=keys
            ver_df[id]=values
    finally:
        bash("rm -rf download/results")            
    return ver_df
        
def get_url_df(operator,header):
    scenario_url="//10.72.1.239/Dumps/{}/".format(operator['scenario'])
    result_url="//10.72.1.239/Results/{job}/{id}".format(job=operator['result'],id=header[1])
    allure_url="http://10.72.1.46/view/{operator}/job/{job}/{id}/allure/".format(operator=operator['provider'],job=operator['job'],id=header[1])

    keys=['Логи и конфиги','Pcap+сценарий(ini)', 'Allure-report']
    values=[result_url,scenario_url,allure_url]

    url_df=pd.DataFrame(
                {' ': keys,
                 'url': values
                	 })
    return url_df

def get_skipped_df(url_with_dict):
    def get_skipped_dict(url_list):
        for url in url_list:
            bash(url)
        grep = bash('grep -r "xfail" download/').value()
        grep_list=grep.split("\n")
        grep_dict={}
        for item in grep_list[1:]:
            test_name=item[item.find("/")+1:item.find(":")-4].strip()
            test_name=test_name[test_name.find("/")+1:]
            reason=item[item.find(":")+7:].strip().strip("'")
            grep_dict[test_name]=reason
        return grep_dict

    wget_cmd='wget -nd -r -l2 --no-parent -A ".ini" --directory-prefix=./download/scenario ftp://10.72.1.239/Dumps/{}/'
    url_list=[wget_cmd.format(operator['scenario']) for operator in url_with_dict]

    skipped_dict=get_skipped_dict(url_list)
    keys=list(skipped_dict.keys())
    values=list(skipped_dict.values())
    
    skipped_df=pd.DataFrame(
        {'TEST': keys,
         'Reason of FAIL': values
        })
    return skipped_df

def sendmail(fromaddr,toaddr,subj,body,filename,att_path):   
    msg = MIMEMultipart()
     
    msg['From'] = fromaddr
    msg['To'] = toaddr
    msg['Subject'] = subj
     
    msg.attach(MIMEText(body, 'plain'))
    attachment = open(att_path, "rb")
     
    part = MIMEBase('application', 'octet-stream')
    part.set_payload((attachment).read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', "attachment; filename= %s" % filename)
     
    msg.attach(part)
     
    server = smtplib.SMTP('10.72.1.210:25')
    server.starttls()
    text = msg.as_string()
    server.sendmail(fromaddr, toaddr, text)
    server.quit()

