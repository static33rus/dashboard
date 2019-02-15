import os
import smtplib
import time
import pandas as pd
import openpyxl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from bash import bash
from redminelib import Redmine
from openpyxl.styles import Font, Border, Side, Style, Color, PatternFill, Alignment
from openpyxl.utils import coordinate_from_string

class Report:
	def __init__(self,):
		self.total_df=pd.DataFrame(columns=[" ","PASSED","FAILED","SKIPPED"])
		self.total_previous_df=pd.DataFrame(columns=[" ","PASSED","FAILED","SKIPPED"])
		self.diff_table=pd.DataFrame(columns=['Стали работать','Перестали работать','Перестали запускаться'])
		self.total_passed,self.total_failed,self.total_skipped=0,0,0
		self.previous_passed,self.previous_failed,self.previous_skipped=0,0,0
		self.redmine = Redmine('https://pm.osnovalab.ru', key='d100e8f8053fc582337cca077b75569f759499c2',requests={'verify': False})

	def create_diff_table(self,df,provider,builds_to_diff="default",include_provider_name=False):
		if builds_to_diff=="default":
			if len(df.columns)>2:
				build1=1
				build2=2
			else:
				build1=build2=1
		else:
			try:
				build1=list(df.columns).index(str(builds_to_diff[0]))
				build2=list(df.columns).index(str(builds_to_diff[1]))
			except ValueError:
				print("В отчете нет информации по сборкам, которые хотите сравнить. Провайдер: {}".format(provider))
				empty_df=pd.DataFrame({'A' : []})
				return empty_df
		test=df[df.columns[0]]
		today=df[df.columns[build1]]
		yesterday=df[df.columns[build2]]
		
		if include_provider_name:
			passed=[" "]
			failed=[provider]
			norun=[" "]
		else:
			passed=[]
			failed=[]
			norun=[]        
		for i in range(0,len(test)):
			if str(today[i])==str(yesterday[i]):
				continue
			if str(today[i])=='PASSED' and str(yesterday)!="PASSED":
				passed.append(test[i])
			if str(today[i])=='FAILURE' and (str(yesterday)!='FAILURE' or str(yesterday)=='SKIPPED'):
				failed.append(test[i])
			if str(today[i])=='' and str(yesterday)!='':
				norun.append(test[i])
		
		s1=pd.Series(passed,name='Стали работать')
		s2=pd.Series(failed,name='Перестали работать')
		s3=pd.Series(norun,name='Перестали запускаться')
			
		df1=pd.concat([s1,s2,s3], axis=1)
		self.diff_table=self.diff_table.append(df1)
		return df1

	def add_summary_to_df(self,df,operator):
		index=2 if len(df.columns)>2 else 1
		passed=self.countif_in_rows("PASSED",df)
		failed=self.countif_in_rows("FAILURE",df)
		skipped=self.countif_in_rows("SKIPPED",df)
		results=["SUMMARY:"]
		for i in range (0,len(passed)-1):
			results.append(" ")
		df.loc[len(df)]=results
		df.loc[len(df)]=passed
		df.loc[len(df)]=failed
		df.loc[len(df)]=skipped
		self.total_passed,self.total_failed,self.total_skipped=self.total_passed+passed[1],self.total_failed+failed[1],self.total_skipped+skipped[1]
		self.previous_passed,self.previous_failed,self.previous_skipped=self.previous_passed+passed[index],self.previous_failed+failed[index],self.previous_skipped+skipped[index]
		self.total_df=self.total_df.append(pd.DataFrame({
								  ' ': operator,
								  'PASSED': passed[1],
								  'FAILED': failed[1],
								  'SKIPPED': skipped[1]}, index=[0]), sort=False)
		self.total_previous_df=self.total_previous_df.append(pd.DataFrame({
								  ' ': operator,
								  'PASSED': passed[index],
								  'FAILED': failed[index],
								  'SKIPPED': skipped[index]}, index=[0]), sort=False)
		return df

	def countif_in_rows(self,word,df):
		##count in each columns by word, return list of counts
		x=['TOTAL '+word]
		n=0
		for column in df.columns[1:]:
			for cell in df[column]:
				if cell==word:
					n=n+1
			x.append(n)
			n=0
		return(x)

	def get_total_tables(self):
		self.total_df.loc[len(self.total_df)]=["Всего: ",self.total_passed,self.total_failed,self.total_skipped]
		self.total_previous_df.loc[len(self.total_df)]=["Всего: ",self.previous_passed,self.previous_failed,self.previous_skipped]
		return self.total_df, self.total_previous_df, self.diff_table

	def sendmail(self,fromaddr,toaddr,subj,body,filename,att_path):   
		msg = MIMEMultipart()
		 
		msg['From'] = fromaddr
		msg['To'] = toaddr
		msg['Subject'] = subj
		 
		msg.attach(MIMEText(body, 'plain'))
		attachment = open(att_path, "rb")
		 
		part = MIMEBase('application', 'octet-stream')
		part.set_payload((attachment).read())
		encoders.encode_base64(part)
		part.add_header('Content-Disposition', "attachment; filename= %s" % filename)
		 
		msg.attach(part)
		 
		server = smtplib.SMTP('10.72.1.210:25')
		server.starttls()
		text = msg.as_string()
		server.sendmail(fromaddr, toaddr, text)
		server.quit()

	def find_in_table_and_get_row_index(self,df,col,findlist):
		index=[]
		for item in findlist:
			index=index+df.index[df[col] == item].tolist()
		return index

	def get_issues_from_redmine(self,search='[autotest]'):
		issues=list(self.redmine.issue.search(search, titles_only=True, open_issues=True))
		url=[issue.url for issue in issues]
		title=[issue.title.split(":")[1] for issue in issues]
		status=[issue.title.split(":")[0][issue.title.split(":")[0].find("("):].strip("(").strip(")") for issue in issues]
		s1=pd.Series(title,name='Issue')
		s2=pd.Series(url,name='url')
		s3=pd.Series(status,name='Status')
		df1=pd.concat([s1,s2,s3], axis=1)
		return df1
	


class Excel_formatting:
	def __init__(self,path):
		self.wb = openpyxl.load_workbook(path)
		self.border = Border(left=Side(border_style='thin', color='000000'),
					right=Side(border_style='thin', color='000000'),
					top=Side(border_style='thin', color='000000'),
					bottom=Side(border_style='thin', color='000000'))

	def auto_column_width(self,ws):
		self.worksheet=self.wb[ws]
		for col in self.worksheet.columns:
			 max_length = 5
			 column = col[0].column # Get the column name
			 for cell in col:
				 try: # Necessary to avoid error on empty cells
					 if len(str(cell.value)) > max_length:
						 max_length = len(cell.value)
				 except:
					 pass
			 adjusted_width = (max_length + 3)
			 self.worksheet.column_dimensions[column].width = adjusted_width

	def conditional_formatting(self,ws,conditional_format_dict):
		cond_list=list(conditional_format_dict.items())
		self.worksheet=self.wb[ws]
		for col in self.worksheet.columns:
			for cell in col:
				try: # Necessary to avoid error on empty cells
					for condition in cond_list:
						if str(cell.value) == condition[0]:
							cell.style=Style(fill=PatternFill(patternType='solid',
												fill_type='solid', 
												fgColor=Color(condition[1])))
							cell.border = self.border
				except:
					pass

	def num_to_letter(self,n):
		###Example of this function: input=6, output=F
		string = ""
		while n > 0:
			n, remainder = divmod(n - 1, 26)
			string = chr(65 + remainder) + string
		return string

	def set_border(self,ws, cell_range,color="DDD9C4"):
		tmp=cell_range.split(":")
		a=tmp[0][1:]
		b=tmp[1][0]+a
		header=tmp[0]+":"+b
	
		rows = ws[cell_range]
		for row in rows:
			for cell in row:
				cell.border = self.border
	
		head=ws[header]
		for row in head:
			for cell in row:
				cell.style=Style(fill=PatternFill(patternType='solid',
										fill_type='solid', 
										fgColor=Color(color)))
				cell.border = self.border
				cell.alignment = Alignment(horizontal='center')

	def fill_row(self,ws,cell_range,color="DDD9C4",set_border=True,alignment='center'):
		rows = ws[cell_range]
		for row in rows:
			for cell in row:
				cell.style=Style(fill=PatternFill(patternType='solid',
										fill_type='solid', 
										fgColor=Color(color)))
				cell.border = self.border if set_border==True else print(".")
				cell.alignment = Alignment(horizontal=alignment)

	def format_table(self,df_shape,start_col=1,start_row=1,fill_row=None,description=None,fill_operator=None):
		rows, columns = df_shape
		row_ident=start_row-1
		col_ident=start_col-1
		start=self.num_to_letter(start_col)+str(start_row)
		end=self.num_to_letter(columns+col_ident)+str(rows+row_ident+1)
		self.set_border(self.worksheet, "{}:{}".format(start,end))
		row_end=rows+row_ident+2
### Add description
		if description is not None:
			merge_cells="{}:{}".format(self.num_to_letter(start_col)+str(start_row-1),self.num_to_letter(columns+col_ident)+str(start_row-1))
			self.worksheet.merge_cells(merge_cells)
			self.worksheet.cell(row=start_row-1, column=start_col).font = Font(color="FF0000", bold=True)
			self.worksheet.cell(row=start_row-1, column=start_col).value = description
			self.worksheet.cell(row=start_row-1, column=start_col).alignment=Alignment(horizontal="center")
### Подкрасим строчки по запоросу пользователя
		if fill_row=="last":
			self.fill_row(self.worksheet, "{}:{}".format(self.num_to_letter(start_col)+str(rows+row_ident+1),self.num_to_letter(columns+col_ident)+str(rows+row_ident+1)),color="95B3D7",alignment="right")
		if fill_row=="summary":
		###Захардкожено подкрасить последние 4 строки, если менять то в str(rows+row_ident-2). -2 отвечает за диапазон строк
			self.fill_row(self.worksheet, "{}:{}".format(self.num_to_letter(start_col)+str(rows+row_ident-2),self.num_to_letter(columns+col_ident)+str(rows+row_ident+1)),color="95B3D7",alignment="left")
		if fill_row=="second":
			self.fill_row(self.worksheet, "{}:{}".format(self.num_to_letter(start_col)+str(2+row_ident),self.num_to_letter(columns+col_ident)+str(2+row_ident)),color="95B3D7")
		if fill_operator is not None:
			for item in fill_operator:
				self.fill_row(self.worksheet, "{}:{}".format(self.num_to_letter(start_col)+str(item+row_ident+2),self.num_to_letter(columns+col_ident)+str(item+row_ident+2)),color="95B3D7")
		return row_end,columns+1

	def insert_image(self,ws,col_end,filepath):
		img = openpyxl.drawing.Image(filepath)
		img.anchor(self.wb[ws].cell('{}1'.format(self.num_to_letter(col_end+10))))
		self.wb[ws].add_image(img)

	def save(self,path):
		self.wb.save(path)

