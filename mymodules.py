import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from bash import bash
import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, Style, Color, PatternFill
import xlsxwriter
import os
import time
from selenium import webdriver

PATH=os.getcwd()

def sendmail(fromaddr,toaddr,subj,body,filename,att_path):	 
	msg = MIMEMultipart()
	 
	msg['From'] = fromaddr
	msg['To'] = toaddr
	msg['Subject'] = subj
	 
	msg.attach(MIMEText(body, 'plain'))
	attachment = open(att_path, "rb")
	 
	part = MIMEBase('application', 'octet-stream')
	part.set_payload((attachment).read())
	encoders.encode_base64(part)
	part.add_header('Content-Disposition', "attachment; filename= %s" % filename)
	 
	msg.attach(part)
	 
	server = smtplib.SMTP('10.72.1.210:25')
	server.starttls()
	text = msg.as_string()
	server.sendmail(fromaddr, toaddr, text)
	server.quit()

def skipped_dict(url_list):
    for url in url_list:
        bash(url)
    grep = bash('grep -r "xfail" download/').value()
    grep_list=grep.split("\n")
    grep_dict={}
    for item in grep_list[1:]:
        test_name=item[item.find("/")+1:item.find(":")-4].strip()
        test_name=test_name[test_name.find("/")+1:]
        reason=item[item.find(":")+7:].strip().strip("'")
        grep_dict[test_name]=reason
    return grep_dict

def set_border(ws, cell_range, fill=False, color="DDD9C4"):
    border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))
    if fill==False:
        tmp=cell_range.split(":")
        a=tmp[0][1:]
        b=tmp[1][0]+a
        header=tmp[0]+":"+b
    
        rows = ws[cell_range]
        for row in rows:
            for cell in row:
                cell.border = border
    
        head=ws[header]
        for row in head:
            for cell in row:
                cell.style=Style(fill=PatternFill(patternType='solid',
                                        fill_type='solid', 
                                        fgColor=Color(color)))
                cell.border = border
    if fill==True:
        rows = ws[cell_range]
        for row in rows:
            for cell in row:
                cell.style=Style(fill=PatternFill(patternType='solid',
                                        fill_type='solid', 
                                        fgColor=Color(color)))
                cell.border = border

def remove_rows_csv(df, word='stroka_kotoroy_ne_doljno_bit_nikogda'):
    x=[]
    for i in range(0,len(df)):
        n=0
        for item in df.loc[i]:
            if type(item)==float:
                n=n+1
            if type(item)==str and item.count(word):
                x.append(i)  
        if n==5:
            x.append(i)
    df1=df.drop(x, axis=0)
    return df1


def download_from_url(url_list,button_id):
    options = webdriver.ChromeOptions()
    options.headless = True
    options.add_experimental_option("prefs", {
      "download.default_directory": PATH+"/download/",
      "download.prompt_for_download": False,
      "download.directory_upgrade": True,
      "safebrowsing.enabled": True
    })
    n=len(url_list)
    for url in url_list:
        newname="download/Test Results "+"("+str(n)+")"+".csv"
        driver = webdriver.Chrome(PATH+'/chromedriver',chrome_options=options)  # Optional argument, if not specified will search path.
        driver.command_executor._commands["send_command"] = ("POST", '/session/$sessionId/chromium/send_command')
        params = {'cmd': 'Page.setDownloadBehavior', 'params': {'behavior': 'allow', 'downloadPath': "/home/m_pavlov/Desktop/allure-report/download/"}}
        command_result = driver.execute("send_command", params)
        driver.get(url);
        search_box = driver.find_element_by_id(button_id).click()
        time.sleep(1)
        os.rename('download/Test Results.csv', newname)
        n-=1
    time.sleep(2)
    return True


def countif_in_rows(word,csv):
##Open csv file, count in each columns by word, return list of counts and csv file
    df=pd.read_csv(csv)
    x=['TOTAL '+word]
    n=0
    for column in df.columns[1:]:
        for cell in df[column]:
            if cell==word:
                n=n+1
        x.append(n)
        n=0
    return(x,df)

def diff_table(csv):
    df=pd.read_csv(csv)
    test=df[df.columns[0]]
    today=df[df.columns[1]]
    yesterday=df[df.columns[2]]
    
    passed=[]
    failed=[]
    norun=[]
    for i in range(0,len(test)-1):
    	if str(today[i])==str(yesterday[i]):
    		continue
    	if str(today[i])=='PASSED' and str(yesterday)!="PASSED":
    		passed.append(test[i])
    	if str(today[i])=='FAILED' and (str(yesterday)!='FAILED' or str(yesterday)=='SKIPPED'):
    		failed.append(test[i])
    	if str(today[i])=='nan' and str(yesterday)!='nan':
    		norun.append(test[i])
    
    
    s1=pd.Series(passed,name='Стали работать')
    s2=pd.Series(failed,name='Перестали работать')
    s3=pd.Series(norun,name='Перестали запускаться')
    
    
    df1=pd.concat([s1,s2,s3], axis=1)
    return df1