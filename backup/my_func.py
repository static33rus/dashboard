import os
import time
from selenium import webdriver
import pandas as pd
import xlsxwriter
import time
import openpyxl
from openpyxl.styles import Border, Side, Style, Color, PatternFill

PATH=os.getcwd()

def get_url(operators):    
    for operator in operators:
        download_url=[]
        provider=operator['provider']
        dvo_list=operator['dvo']
        job=operator['job']
        for dvo in dvo_list:
            download_url.append("http://10.72.1.46/view/{provider}/job/{job}/dumptemplate={dvo},label=SORM_server/test_results_analyzer/".format(provider=provider,job=job, dvo=dvo))
        operator['url']=download_url
    return operators

def download_from_url(url_list,num,sleep=3):
    options = webdriver.ChromeOptions()
    options.headless = False
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    driver = webdriver.Chrome(PATH+'/chromedriver',chrome_options=options)  # Optional argument, if not specified will search path.
    driver.command_executor._commands["send_command"] = ("POST", '/session/$sessionId/chromium/send_command')
    params = {'cmd': 'Page.setDownloadBehavior', 'params': {'behavior': 'allow', 'downloadPath': PATH+"/download/"}}
    command_result = driver.execute("send_command", params)
    for operator in url_list:
        n=len(operator['url'])
        for url in operator['url']:
            newname="download/{}{}.csv".format(operator['provider'],str(n))
            driver.get(url);
            if num!=5:
                search_box = driver.find_element_by_id("settingsmenubutton").click()
                driver.find_element_by_id("noofbuilds").clear()
                driver.find_element_by_id("noofbuilds").send_keys(str(num))
                time.sleep(sleep)
            search_box = driver.find_element_by_id('downloadCSV').click()
            time.sleep(1)
            os.rename('download/Test Results.csv', newname)
            n-=1
    time.sleep(2)
    return True

def remove_rows_csv(df,num,word="Esli eto slovo est v stroke, to stroka udalit'sya"):
    #if row will contain word, row will be deleted
    x=[]
    for index, row in df.iterrows():
        n=0
        for item in row:
            if type(item)==float:
                n+=1
            if type(item)==str:
                if item.count(word)>0:
                    x.append(index)
                    continue
        if n==num:
            x.append(index)
    df1=df.drop(x, axis=0)
    return df1

def colnum_string(n):
    ###Example of this function: input=6, output=F
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

def merge_df_and_save_to_excel(dict_with_url, number_of_progons):
    length_dfs=[]
    cols=list(range(2,number_of_progons+3))
    timestr = time.strftime("%Y.%m.%d")
    excel_writer = pd.ExcelWriter("report/"+timestr+".xlsx", engine='xlsxwriter')
    for operator in dict_with_url:
        df = pd.read_csv("download/{}1.csv".format(operator['provider']),usecols=cols)
        df = remove_rows_csv(df,number_of_progons,"SORM.test")
        header=list(df.columns)
        finaldf=pd.DataFrame(columns=header)
        finaldf=finaldf.append(df,sort=False)
        for i in range (2,len(operator['url'])+1):
            df = pd.read_csv("download/{}{}.csv".format(operator['provider'],i),usecols=cols)
            df = remove_rows_csv(df,number_of_progons,"SORM.test")
            finaldf=finaldf.append(df,sort=False)
        length_dfs.append(len(finaldf)+1)

        finaldf.to_excel(excel_writer, sheet_name=operator['provider'],index=False)
        workbook  = excel_writer.book
        worksheet = excel_writer.sheets[operator['provider']]
        red = workbook.add_format({'bg_color': '#FFC7CE',
                                   'font_color': '#9C0006'})
        green = workbook.add_format({'bg_color': '#81F781',
                                       'font_color': '#9C0006'})
        header_format = workbook.add_format({
                                     'bold': True,
                                     'text_wrap': True,
                                     'valign': 'top',
                                     'fg_color': '#DDD9C4',
                                     'border': 1})
        width = 40
        
        worksheet.set_column(0, 0, width)
        worksheet.set_column(number_of_progons+2, number_of_progons+2, width/2)
        worksheet.set_column(number_of_progons+3, number_of_progons+3, width+50)
        end=colnum_string(number_of_progons+1)+str(len(finaldf)+1)
        worksheet.conditional_format('B2:'+end, {'type':     'cell',
                                                 'criteria': '==',
                                                 'value':    '"FAILED"',
                                                 'format':   red})
    
    
        worksheet.conditional_format('B2:'+end, {'type':     'cell',
                                                 'criteria': '==',
                                                 'value':    '"PASSED"',
                                                 'format':   green})

    excel_writer.save()
    return timestr, length_dfs

def set_border(ws, cell_range, fill=False, color="DDD9C4"):
    border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))
    if fill==False:
        tmp=cell_range.split(":")
        a=tmp[0][1:]
        b=tmp[1][0]+a
        header=tmp[0]+":"+b
    
        rows = ws[cell_range]
        for row in rows:
            for cell in row:
                cell.border = border
    
        head=ws[header]
        for row in head:
            for cell in row:
                cell.style=Style(fill=PatternFill(patternType='solid',
                                        fill_type='solid', 
                                        fgColor=Color(color)))
                cell.border = border
    if fill==True:
        rows = ws[cell_range]
        for row in rows:
            for cell in row:
                cell.style=Style(fill=PatternFill(patternType='solid',
                                        fill_type='solid', 
                                        fgColor=Color(color)))
                cell.border = border